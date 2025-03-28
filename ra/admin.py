from django.contrib import admin
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.http import HttpResponse
from datetime import date
from ra.models import Avaliacao, Criterio,Colaborador

# Register your models here.

@admin.register(Colaborador)
class ColaboradorAdmin(admin.ModelAdmin):
    list_display =('nome',)

@admin.register(Criterio)
class CriterioAdmin(admin.ModelAdmin):
    list_display=('nome',)

@admin.register(Avaliacao)
class AvaliacaoAdmin(admin.ModelAdmin):
    list_display=('colaborador','criterio', 'nota',)
    
    #importando para excel
    def export_product_to_excel(request,self,queryset):
    # Cria o workbook e a planilha
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Avaliação R.A"

        # Adiciona o cabeçalho
        headers = ['Nome', 'Criterio','Nota',]
        worksheet.append(headers)
        
        #Estilizacao cabecalho
        # header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Amarelo
        # header_font = Font(bold=True, color="000000")  # Texto preto em negrito

        # for col_num, cell in enumerate(worksheet[1], 1):
        #     cell.fill = header_fill
        #     cell.font = header_font
        
        # Recupera os dados do modelo e preenche a planilha
        avaliacao = Avaliacao.objects.all() #Busca Avaliação todos o objetos
        data_atual= date.today()
        
        for computers in avaliacao: #Percorre os objetos
            worksheet.append([
                str (computers.colaborador),
                str (computers.criterio),
                int (computers.nota),
            #   controle.delivery.strftime("%Y-%m-%d"),
            #   controle.delivery.strftime("%H:%M:%S"),
            ])
        
        # Criando uma tabela
        last_row = len(avaliacao) + 1  # Define o tamanho da tabela
        tabela = Table(displayName="TabelaAvaliacao", ref=f"A1:C{last_row}")

        # Estilizando a tabela
        estilo_tabela = TableStyleInfo(
            name="TableStyleMedium14",  # Escolha um estilo disponível no Excel
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,  # Listras nas linhas
            showColumnStripes=False
        )
        tabela.tableStyleInfo = estilo_tabela

        # Adiciona a tabela à planilha
        worksheet.add_table(tabela)
        
        # Centralizar todas as colunas da tabela
        for row in worksheet.iter_rows(min_row=1, max_row=last_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Configura a resposta HTTP para o download
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename=Avaliacao_RA_{data_atual}.xlsx"
        workbook.save(response)
        return response
    
    export_product_to_excel.short_description = 'Exportar para excel'
    actions = [export_product_to_excel]
    
    
    

    
